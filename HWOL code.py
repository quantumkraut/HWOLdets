import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import tempfile
import re

# -------------------------------------------------
# APP HEADER
# -------------------------------------------------
st.set_page_config(page_title="HWOL OEMA (DETS)")
st.title("HWOL OEMA (DETS)")

# -------------------------------------------------
# FILE UPLOAD
# -------------------------------------------------
labfile = st.file_uploader(
    "Upload OEMA (DETS) lab Excel File (.xlsx)",
    type=["xlsx"]
)

columnno = st.text_input(
    "Required sample column number (only change if more than one sample):",
    value="1"
)

process = st.button("Click to Generate OEMA HWOL")

status = st.empty()

# -------------------------------------------------
# HELPER FUNCTIONS (same logic as R)
# -------------------------------------------------
def safe_grep(pattern, series):
    matches = series[series.str.contains(pattern, regex=True, na=False)]
    return matches.iloc[0] if not matches.empty else ""

def safe_add(x, y):
    x, y = str(x), str(y)
    has_lt = x.startswith("<") or y.startswith("<")

    x_num = float(re.sub(r"[^0-9eE\.-]", "", x)) if x else 0
    y_num = float(re.sub(r"[^0-9eE\.-]", "", y)) if y else 0

    total = x_num + y_num
    return f"< {total}" if has_lt else total

def convert_ug_to_mg(x):
    x = str(x)
    has_lt = x.startswith("<")
    num = float(re.sub(r"[^0-9eE\.-]", "", x))
    num = num / 1000
    return f"< {num}" if has_lt else num

# -------------------------------------------------
# PROCESS BUTTON
# -------------------------------------------------
if process:

    if labfile is None:
        status.error("Please upload a file first.")
        st.stop()

    try:
        status.info("Processing file...")

        # ---- LOAD TEMPLATE (pandas version of read_xlsx) ----
        template = pd.read_excel("MultiSamplesOEMA.xlsx")
        template = template.iloc[:, :3]

        # ---- LOAD LAB FILE ----
        test = pd.read_excel(labfile)
        col_index = 4 + int(columnno)

        test = test.iloc[16:, [0, col_index]]
        test.columns = ["Determinant", "Result"]
        test.reset_index(drop=True, inplace=True)

        # ---- ROUND DECIMALS (Python equivalent of sapply) ----
        def round_values(x):
            if pd.isna(x):
                return x
            x = str(x)
            if not re.match(r"^[0-9.\-]+$", x):
                return x
            if "." in x and len(x.split(".")[1]) >= 4:
                return f"{round(float(x), 2):.2f}"
            return x

        test["Result"] = test["Result"].apply(round_values)

        # ---- BUILD ORDERED TABLE (same structure as R) ----
        test2 = pd.DataFrame({
            "Determinant": [
                "",
                "",
                safe_grep("Antimony", test["Determinant"]),
                safe_grep("Arsenic", test["Determinant"]),
                safe_grep("Barium", test["Determinant"]),
                safe_grep("Beryllium", test["Determinant"]),
                safe_grep("Boron", test["Determinant"]),
                safe_grep("Cadmium", test["Determinant"]),
                safe_grep("Chromium \\(aqua", test["Determinant"]),
                safe_grep("hexavalent", test["Determinant"]),
                safe_grep("Copper", test["Determinant"]),
                safe_grep("Cobalt", test["Determinant"]),
                safe_grep("Lead", test["Determinant"]),
                safe_grep("Manganese", test["Determinant"]),
                safe_grep("Mercury", test["Determinant"]),
                safe_grep("Molybdenum", test["Determinant"]),
                safe_grep("Nickel", test["Determinant"]),
                safe_grep("Selenium", test["Determinant"]),
                safe_grep("Thallium", test["Determinant"]),
                safe_grep("Tin", test["Determinant"]),
                safe_grep("Calcium", test["Determinant"]),
                safe_grep("Phophorous", test["Determinant"]),
                safe_grep("Vanadium", test["Determinant"]),
                safe_grep("Sulphur", test["Determinant"]),
                safe_grep("Zinc", test["Determinant"]),
                safe_grep("Iron", test["Determinant"]),
                safe_grep("TPH Total", test["Determinant"]),
                "",
                safe_grep("MTBE", test["Determinant"]),
                safe_grep("Benzene", test["Determinant"]),
                safe_grep("Toluene", test["Determinant"]),
                safe_grep("Ethylbenzene", test["Determinant"]),
                "Total Xylene"
            ]
        })

        joined = test2.merge(test, how="left", on="Determinant")

        # ---- XYLENE ADDITION ----
        joined.loc[32, "Result"] = safe_add(
            test.loc[test["Determinant"] == "p & m-Xylene", "Result"].values[0],
            test.loc[test["Determinant"] == "o-Xylene", "Result"].values[0]
        )

        # ---- UNIT CONVERSION ----
        joined.loc[28:32, "Result"] = joined.loc[28:32, "Result"].apply(convert_ug_to_mg)

        # ---- WRITE TO EXCEL TEMPLATE ----
        wb = load_workbook("MultiSamplesOEMA.xlsx")
        ws = wb.active

        for i, value in enumerate(joined["Result"], start=3):
            ws.cell(row=i, column=3).value = value

        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        wb.save(tmp.name)

        status.success("Done! Download below.")
        st.download_button(
            "Download HWOL",
            data=open(tmp.name, "rb"),
            file_name=f"OEMA_HWOL_{pd.Timestamp.today().date()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        status.error(str(e))

